#!/usr/bin/env python3
"""
One-shot converter: xlsx -> Univer IWorkbookData JSON.

Reads the kosztorys xlsx and emits a JSON file the Univer spike route loads
at runtime via fetch().

Fidelity:
- values + formulas (cross-sheet refs preserved verbatim)
- number formats
- font color, bg fill, bold/italic/underline, font size, font family
- borders (t/b/l/r) mapped to Univer BorderStyleType enum
- horizontal + vertical alignment, wrap
- column widths, row heights
- merged cells
- frozen panes
- tab order

Limitations:
- conditional formatting, data validations, charts, images: skipped
- theme-color resolution: best-effort, only direct rgb is read; theme refs become None
- Excel column width (char units) -> Univer pixel width: width * 7 + 5
- Excel row height (points) -> Univer pixel height: points * 4/3
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string


# --- constants ---

XLSX_PATH = Path("Kopia kosztorys wzór dla konrada do testów.xlsx")
OUT_PATH = Path("public/data/kosztorys-workbook.json")

# openpyxl border style name -> Univer BorderStyleType
# https://reference.univer.ai/en-US/classes/IStyleData
BORDER_STYLE_MAP = {
    None: 0,            # NONE
    "thin": 1,          # THIN
    "hair": 2,          # HAIR
    "dotted": 3,        # DOTTED
    "dashed": 4,        # DASHED
    "dashDot": 5,       # DASH_DOT
    "dashDotDot": 6,    # DASH_DOT_DOT
    "double": 7,        # DOUBLE
    "medium": 8,        # MEDIUM
    "mediumDashed": 9,  # MEDIUM_DASHED
    "mediumDashDot": 10,
    "mediumDashDotDot": 11,
    "slantDashDot": 12,
    "thick": 13,        # THICK
}

# Excel HorizontalAlignment -> Univer HorizontalAlign (0=left, 1=center, 2=right, 3=justify)
H_ALIGN_MAP = {
    "left": 1,
    "center": 2,
    "right": 3,
    "justify": 4,
    "centerContinuous": 2,
    "distributed": 4,
    "fill": 1,
    "general": 0,
}

# VerticalAlign (0=top, 1=middle, 2=bottom)
V_ALIGN_MAP = {
    "top": 1,
    "center": 2,
    "bottom": 3,
    "justify": 2,
    "distributed": 2,
}


# --- helpers ---


def col_width_px(width_chars: float | None, default: float = 8.43) -> int:
    """Excel char-width -> Univer px (approx for Calibri 11)."""
    w = width_chars if width_chars is not None else default
    return max(20, round(w * 7 + 5))


def row_height_px(height_points: float | None, default: float = 15.0) -> int:
    """Excel points -> Univer px."""
    h = height_points if height_points is not None else default
    return max(15, round(h * 4 / 3))


def rgb_of(color: Any) -> str | None:
    """Extract '#RRGGBB' from an openpyxl Color, or None for theme refs we can't resolve."""
    if color is None:
        return None
    try:
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb
            # openpyxl gives 'AARRGGBB' or 'RRGGBB'
            if isinstance(rgb, str):
                if len(rgb) == 8:
                    return f"#{rgb[2:].upper()}"
                if len(rgb) == 6:
                    return f"#{rgb.upper()}"
    except (AttributeError, ValueError):
        return None
    return None


def style_signature(style: dict[str, Any]) -> str:
    """Stable hash for style dedup."""
    return json.dumps(style, sort_keys=True, separators=(",", ":"))


class StyleRegistry:
    def __init__(self) -> None:
        self._by_sig: dict[str, str] = {}
        self.styles: dict[str, dict[str, Any]] = {}
        self._counter = 0

    def intern(self, style: dict[str, Any]) -> str | None:
        if not style:
            return None
        sig = style_signature(style)
        if sig in self._by_sig:
            return self._by_sig[sig]
        self._counter += 1
        sid = f"s{self._counter}"
        self._by_sig[sig] = sid
        self.styles[sid] = style
        return sid


def build_cell_style(cell: Cell) -> dict[str, Any]:
    style: dict[str, Any] = {}
    font = cell.font
    fill = cell.fill
    border = cell.border
    align = cell.alignment
    nfmt = cell.number_format

    # font
    if font is not None:
        if font.bold:
            style["bl"] = 1
        if font.italic:
            style["it"] = 1
        if font.underline and font.underline != "none":
            style["ul"] = {"s": 1}
        size = font.sz
        if size is not None and size != 11:
            style["fs"] = float(size)
        if font.name and font.name != "Calibri":
            style["ff"] = font.name
        font_rgb = rgb_of(font.color)
        if font_rgb and font_rgb != "#000000":
            style["cl"] = {"rgb": font_rgb}

    # fill (background)
    if fill is not None and fill.patternType == "solid":
        bg_rgb = rgb_of(fill.fgColor) or rgb_of(fill.start_color)
        if bg_rgb and bg_rgb != "#FFFFFF":
            style["bg"] = {"rgb": bg_rgb}

    # borders
    bd: dict[str, Any] = {}
    for key, side in (("t", border.top), ("b", border.bottom), ("l", border.left), ("r", border.right)):
        if side is None:
            continue
        s_num = BORDER_STYLE_MAP.get(side.style, 0)
        if s_num == 0:
            continue
        side_rgb = rgb_of(side.color) or "#000000"
        bd[key] = {"s": s_num, "cl": {"rgb": side_rgb}}
    if bd:
        style["bd"] = bd

    # alignment
    if align is not None:
        if align.horizontal and align.horizontal in H_ALIGN_MAP:
            style["ht"] = H_ALIGN_MAP[align.horizontal]
        if align.vertical and align.vertical in V_ALIGN_MAP:
            style["vt"] = V_ALIGN_MAP[align.vertical]
        if align.wrap_text:
            style["tb"] = 3  # WRAP

    # number format
    if nfmt and nfmt != "General":
        style["n"] = {"pattern": nfmt}

    return style


def cell_to_univer(cell: Cell, style_id: str | None) -> dict[str, Any] | None:
    """Convert one openpyxl cell to Univer ICellData. Returns None for empty cells with no style."""
    out: dict[str, Any] = {}
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        out["f"] = v
    elif v is not None:
        if isinstance(v, bool):
            out["v"] = 1 if v else 0
        elif isinstance(v, (int, float)):
            out["v"] = v
        else:
            out["v"] = str(v)
    if style_id is not None:
        out["s"] = style_id
    if not out:
        return None
    return out


def sheet_to_univer(ws, registry: StyleRegistry) -> dict[str, Any]:
    cell_data: dict[int, dict[int, Any]] = {}
    formula_count = 0
    cell_count = 0

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            style = build_cell_style(cell)
            style_id = registry.intern(style)
            cell_dict = cell_to_univer(cell, style_id)
            if cell_dict is None:
                continue
            r = cell.row - 1
            c = cell.column - 1
            cell_data.setdefault(r, {})[c] = cell_dict
            cell_count += 1
            if "f" in cell_dict:
                formula_count += 1

    # column widths
    column_data: dict[int, dict[str, Any]] = {}
    for col_letter, dim in ws.column_dimensions.items():
        # dim.min/max can span ranges; key by min..max
        if dim.width is None and not dim.hidden:
            continue
        c_min = dim.min if dim.min else column_index_from_string(col_letter)
        c_max = dim.max if dim.max else c_min
        w = col_width_px(dim.width)
        for c in range(c_min, c_max + 1):
            entry: dict[str, Any] = {"w": w}
            if dim.hidden:
                entry["hd"] = 1
            column_data[c - 1] = entry

    # row heights
    row_data: dict[int, dict[str, Any]] = {}
    for r_idx, dim in ws.row_dimensions.items():
        if dim.height is None and not dim.hidden:
            continue
        h = row_height_px(dim.height)
        entry = {"h": h}
        if dim.hidden:
            entry["hd"] = 1
        row_data[r_idx - 1] = entry

    # merges
    merge_data = []
    for rng in ws.merged_cells.ranges:
        merge_data.append({
            "startRow": rng.min_row - 1,
            "startColumn": rng.min_col - 1,
            "endRow": rng.max_row - 1,
            "endColumn": rng.max_col - 1,
        })

    # freeze panes (openpyxl: 'B2' means freeze cols A and row 1)
    freeze = {"startRow": -1, "startColumn": -1, "xSplit": 0, "ySplit": 0}
    if ws.freeze_panes:
        col_letters = "".join(c for c in ws.freeze_panes if c.isalpha())
        row_str = "".join(c for c in ws.freeze_panes if c.isdigit())
        if col_letters:
            x_split = column_index_from_string(col_letters) - 1
            if x_split > 0:
                freeze["xSplit"] = x_split
                freeze["startColumn"] = x_split
        if row_str:
            y_split = int(row_str) - 1
            if y_split > 0:
                freeze["ySplit"] = y_split
                freeze["startRow"] = y_split

    # default column width
    default_col_w = col_width_px(ws.sheet_format.defaultColWidth) if ws.sheet_format.defaultColWidth else 80
    default_row_h = row_height_px(ws.sheet_format.defaultRowHeight) if ws.sheet_format.defaultRowHeight else 24

    sheet_id = f"sheet-{abs(hash(ws.title)) % (10**10)}"

    print(
        f"  '{ws.title}': {cell_count} cells, {formula_count} formulas, "
        f"{len(merge_data)} merges, {len(row_data)} sized rows, {len(column_data)} sized cols",
        file=sys.stderr,
    )

    return {
        "id": sheet_id,
        "name": ws.title,
        "tabColor": "",
        "hidden": 0,
        "rowCount": max(max_row, 100),
        "columnCount": max(max_col, 20),
        "zoomRatio": 1,
        "freeze": freeze,
        "scrollTop": 0,
        "scrollLeft": 0,
        "defaultColumnWidth": default_col_w,
        "defaultRowHeight": default_row_h,
        "mergeData": merge_data,
        "cellData": cell_data,
        "rowData": row_data,
        "columnData": column_data,
        "showGridlines": 1,
        "rowHeader": {"width": 46, "hidden": 0},
        "columnHeader": {"height": 20, "hidden": 0},
        "rightToLeft": 0,
    }


def main() -> None:
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found (run from repo root).", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {XLSX_PATH}...", file=sys.stderr)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)

    registry = StyleRegistry()
    sheets: dict[str, Any] = {}
    sheet_order: list[str] = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet = sheet_to_univer(ws, registry)
        sheets[sheet["id"]] = sheet
        sheet_order.append(sheet["id"])

    workbook = {
        "id": "kosztorys-workbook",
        "name": "Kosztorys",
        "appVersion": "0.23.0",
        "locale": "plPL",
        "styles": registry.styles,
        "sheetOrder": sheet_order,
        "sheets": sheets,
        "resources": [
            {"name": "SHEET_DEFINED_NAME_PLUGIN", "data": ""},
        ],
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(workbook, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = OUT_PATH.stat().st_size / 1024
    print(
        f"\nWrote {OUT_PATH} ({size_kb:,.0f} KB) — "
        f"{len(sheet_order)} sheets, {len(registry.styles)} unique styles",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
